import openpyxl


wb = openpyxl.Workbook()

ws = wb.create_sheet('hello world', 0)
ws['A2'] = 'hello world'
wb.save('hello.xlsx')