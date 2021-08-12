import time
import random
import openpyxl


def interest():
    wb = openpyxl.load_workbook('information.xlsx')
    ws = wb['credit_card']
    for row in range(1, ws.max_row):
        position = 'F' + str(row)
        if ws[position].value == '待还':
            continue
        ws[position] = int(int(ws[position].value) * 1.002)
        wb.save('information.xlsx')

def is_right_card(card = 0, id = ''):
    if card == 1:
        if (len(id) != 16) or (not id.isdigit()):
            return False
        else:
            return True
    if card == 2:
        if (len(id) != 12) or (not id.isdigit()):
            return False
        else:
            return True

def is_leap_year(year):
    if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0):
        return True
    return False

def is_past(year, month, day):
    today = time.localtime()
    today_year = today.tm_year
    today_month = today.tm_mon
    today_day = today.tm_mday
    if year < today_year:
        return True
    if year == today_year:
        if month < today_month:
            return True
        if month == today_month:
            if day <= today_day:
                return True
            else:
                return False

def right_date(date = ''):
    dates = date.split('-')
    if (len(dates[0]) != 4) or (len(dates[1]) != 2) or (len(dates[2]) != 2):
        return False

    year = int(dates[0])
    month = int(dates[1])
    day = int(dates[2])


    if month == 2:
        if is_leap_year(year):
            if day > 29:
                return False
        else:
            if day > 28:
                return False
    if month in [1, 3, 5, 7, 8, 10, 12]:
        if day > 31:
            return False
    elif month in [4, 6, 9, 11]:
        if day > 30:
            return False
    elif month == 2:
        pass
    else:
        return False

    return is_past(year, month, day)

def create_card():
    id = ''
    secret = ''
    save_credit = input('创建储蓄卡还是信用卡？储蓄卡请输入save，信用卡请输入credit:')
    name = input('姓名:')
    info = []
    if save_credit == 'save':
        while True:
            date = input('请输入生日(yyyy-mm-dd):')
            if right_date(date):
                break
            else:
                print('输入的日期有错误！')

        for each in date.split('-'):
            for i in each:
                info.append(i)

        info.append(str(random.randint(0, 9)))
        info.append(str(random.randint(0, 9)))
        info.append(str(random.randint(0, 9)))
        random.shuffle(info)

        man_woman = input('请输入性别(男/女):')
        if man_woman == '男':
            info.append(random.choice(['1', '3', '5', '7', '9']))
        if man_woman == '女':
            info.append(random.choice(['0', '2', '4', '6', '8']))

        for i in info:
            id += i

        for num in range(6):
            secret += str(random.randint(0, 9))

        print('创建中', end='')
        for i in range(6):
            print('.', end='')
            time.sleep(1)
        print('\n您的账号及密码如下:')
        print('账号:{}'.format(id))
        print('密码:{}'.format(secret))

        need_it = input('是否激活该账号(y/n):')
        if need_it == 'y':
            print('激活成功！')
            return (id, secret, date, man_woman, name, 'save')
        if need_it == 'n':
            print('激活失败！')
            return (False, 0, 0, 0, 0, 0)

    if save_credit == 'credit':
        while True:
            date = input('请输入生日(yyyy-mm-dd):')
            if right_date(date):
                break
            else:
                print('输入的日期有错误！')

        man_woman = input('请输入性别(男/女)')

        for num in range(15):
            id += str(random.randint(0, 9))

        if man_woman == '男':
            id += str(random.choice(['1', '3', '5', '7', '9']))
        if man_woman == '女':
            id += str(random.choice(['0', '2', '4', '6', '8']))

        for num in range(6):
            secret += str(random.randint(0, 9))

        print('创建中', end='')
        for i in range(6):
            print('.', end='')
            time.sleep(1)
        print('\n您的账号及密码如下:')
        print('账号:{}'.format(id))
        print('密码:{}'.format(secret))

        need_it = input('是否激活该账号(y/n):')
        if need_it == 'y':
            print('激活成功！')
            return (id, secret, date, man_woman, name, 'credit')
        if need_it == 'n':
            print('激活失败！')
            return (False, 0, 0, 0, 0, 0)


class Credit_card():
    def __init__(self, card_id, need_pay, most):
        self.card_id = card_id
        self.need_pay = need_pay
        self.most = most
        self.balance = 0

    def repay(self, money_amount = ''):
        if (int(money_amount) <= 0) and (not money_amount.isdigit()):
            print('金额非法！请还款大于零的整数金额！')
        else:
            if self.need_pay < int(money_amount):
                print('不需要还如此多, 还款失败！')
            else:
                self.need_pay -= int(money_amount)
                print('还款成功!')
                print('待还金额如下:{}'.format(self.need_pay))

    def loan(self, money_amount = ''):
        if (int(money_amount) <= 0) and (not money_amount.isdigit()):
            print('金额非法！请选择大于零的整数金额！')
        else:
            if (self.most - self.need_pay) < int(money_amount):
                print('额度不足，请再次尝试！')
            else:
                self.need_pay += int(money_amount)
                self.most -= int(money_amount)
                print('借款成功，待还金额如下:{}'.format(self.need_pay))

    def change_most(self):
        print('低等额度提升:low')
        print('中等额度提升:middle')
        print('高等额度提升:high')
        print('特等额度提升:super')
        print('最高等额度提升:highest')
        print('取消操作请输入cancel')
        witch_change = input('请选择适合的额度:')
        if witch_change == 'low':
            change = random.randint(0, 1000)
        if witch_change == 'middle':
            change = random.randint(1000, 10000)
        if witch_change == 'high':
            change = random.randint(10000, 100000)
        if witch_change == 'super':
            change = random.randint(100000, 1000000)
        if witch_change == 'highest':
            change = random.randint(1000000, 10000000)
        if witch_change == 'cancel':
            return 0
        self.most += change
        print('提高额度成功，目前额度如下{}'.format(self.most))


class Savings_card():
    def __init__(self, card_id, balance):
        self.card_id = card_id
        self.balance = balance

    def deposit(self, money_amount = ''):
        if (int(money_amount) <= 0) and (not money_amount.isdigit()):
            print('存入的金额非法！请存入大于零的整数金额！')
        else:
            self.balance += int(money_amount)
            print('存入中', end='')
            for i in range(6):
                print('.', end='')
                time.sleep(1)
            print('\n成功，余额如下{}:'.format(self.balance))

    def take_the_money(self, money_amount = ''):
        if (int(money_amount) <= 0) and (not money_amount.isdigit()):
            print('金额非法！请取出大于零的整数金额！')
        elif self.balance < int(money_amount):
            print('余额不足，就这点钱还想取这么多出来')
        else:
            self.balance -= int(money_amount)
            print('取钱中', end='')
            for i in range(6):
                print('.', end='')
                time.sleep(1)
            print('\n成功，余额如下{}:'.format(self.balance))

    def transfer(self, target_card, money_amount = ''):
        Y_N = 'n'
        if (self.balance < int(money_amount)) or (not money_amount.isdigit()):
            print('钱不够或者金额非法')
        else:
            Y_N = input('是否要将{}元转入账号{}？(y/n)'.format(money_amount, target_card.card_id))

        if Y_N == 'y':
            self.balance -= int(money_amount)
            target_card.balance += int(money_amount)
            print('操作成功，余额如下:{}'.format(self.balance))


class User():
    def __init__(self, name):
        self.name = name


def main():
    interest()
    wb = openpyxl.load_workbook('information.xlsx')
    while True:
        wb = openpyxl.load_workbook('information.xlsx')
        print('===========================================')
        print('创建卡号请输入create')
        print('读取储蓄卡请输入save_card')
        print('读取信用卡请输入credit_card')
        print('啥都不干请输入quit')
        print('===========================================')

        command = input('做:')
        if command == 'create':
            card_id, secret, birth, gender, name, save_credit = create_card()
            if card_id != False:
                if save_credit == 'save':
                    wb['save_card'].append([name, gender, birth, card_id, secret, 0])
                    wb.save('information.xlsx')
                if save_credit == 'credit':
                    wb['credit_card'].append([name, gender, birth, card_id ,secret, 0, 2000])
                    wb.save('information.xlsx')

        if command == 'save_card':
            codes = 0
            card_id = input('请输入账号：')
            secret = input('请输入密码')
            ws = wb['save_card']
            for row in range(1, ws.max_row + 1):
                position = 'D' + str(row)
                if ws[position].value == card_id:
                    position_1 = 'E' + str(row)
                    if ws[position_1].value == secret:
                        position_2 = 'F' + str(row)
                        save_card = Savings_card(card_id, ws[position_2].value)
                        print('登录成功')
                        codes = 1
                        break
            else:
                print('账号或密码错误！')

            while codes == 1:
                print('+++++++++++++++++++++++++++++++++++')
                print('查看卡信息请输入info')
                print('存款请输入deposit')
                print('取款请输入take_the_money')
                print('转账请输入transfer')
                print('退出请输入exit')
                print('+++++++++++++++++++++++++++++++++++')
                what_do = input('我要做:')
                if what_do == 'info':
                    print('余额:{}'.format(save_card.balance))

                if what_do == 'deposit':
                    money = input('存入多少钱？\n存入：')
                    save_card.deposit(money_amount=money)
                if what_do == 'take_the_money':
                    money = input('取出多少钱？\n取出：')
                    save_card.take_the_money(money_amount=money)

                if what_do == 'transfer':
                    target_id = input('目标账号：')
                    for row in range(1, ws.max_row + 1):
                        position = 'D' + str(row)
                        if ws[position].value == target_id:
                            target_balance_position = 'F' + str(row)
                            target_save = Savings_card(target_id, ws[target_balance_position].value)
                            money = input('转多少钱：')
                            save_card.transfer(target_save, money)

                            for row in range(1, ws.max_row + 1):
                                position = 'D' + str(row)
                                if ws[position].value == target_id:
                                    position_1 = 'F' + str(row)
                                    ws[position_1] = target_save.balance
                            break


                if what_do == 'exit':
                    break

            for row in range(1, ws.max_row + 1):
                position = 'D' + str(row)
                if ws[position].value == card_id:
                    position_1 = 'E' + str(row)
                    if ws[position_1].value == secret:
                        position_2 = 'F' + str(row)
                        ws[position_2] = save_card.balance
                        wb.save('information.xlsx')

        if command == 'credit_card':
            codes = 0
            card_id = input('请输入卡号:')
            secret = input('请输入密码:')

            position_2 = ''
            position_3 = ''
            ws = wb['credit_card']
            for row in range(1, ws.max_row + 1):
                position = 'D' + str(row)
                if ws[position].value == card_id:
                    position_1 = 'E' + str(row)
                    if ws[position_1].value == secret:
                        position_2 = 'F' + str(row)
                        position_3 = 'G' + str(row)
                        credit_card = Credit_card(card_id, ws[position_2].value, ws[position_3].value)
                        print('登录成功')
                        codes = 1
                        break
            else:
                print('账号或密码错误！')

            while codes == 1:
                print('----------------------------------------')
                print('本银行利率为1.002')
                print('查看信息请输入info')
                print('借款请输入loan')
                print('还款请输入repay')
                print('提高额度请输入improve')
                print('退出请输入exit')
                print('----------------------------------------')
                what_do = input('我要做:')
                if what_do == 'info':
                    print('待还金额:{}\n额度:{}'.format(credit_card.need_pay, credit_card.most))

                if what_do == 'loan':
                    much_money = input('需要借款(你的借款额度是{}元)'.format(credit_card.most))
                    credit_card.loan(much_money)

                if what_do == 'repay':
                    much_money = input('需要还款({}待还):'.format(credit_card.need_pay))
                    credit_card.repay(much_money)

                if what_do == 'improve':
                    credit_card.change_most()

                if what_do == 'exit':
                    break

            ws[position_2] = credit_card.need_pay
            ws[position_3] = credit_card.most
            wb.save('information.xlsx')

        if command ==  'quit':
            break

if __name__ == '__main__':
    main()